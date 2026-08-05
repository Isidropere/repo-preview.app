import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Articulos para Intercambio"

    # Enable gridlines
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    even_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    item_name_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    default_font = Font(name="Calibri", size=10, color="334155")
    
    thin_border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Headers
    headers = [
        "Categoria",
        "Articulo / Servicio",
        "Tipo Recomendado",
        "Descripcion / Ejemplo",
        "Valor Estimado (RD$)",
        "Criterios de Intercambio (Ideal por que cambiarlo)"
    ]
    
    ws.append(headers)
    
    # Header styles
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
    
    # Data List
    data = [
        # Instrumento musical
        ("Instrumento musical", "Guitarra Acustica Yamaha F310", "Ambos", "Guitarra clasica ideal para principiantes y nivel medio", "RD$ 8,000 - 12,000", "Cambio por guitarra electrica, teclado electronico o amplificador de similar valor."),
        ("Instrumento musical", "Teclado Electronico Casio CT-S300", "Ambos", "Teclado de 61 teclas con sensibilidad al tacto y conexion USB", "RD$ 10,000 - 15,000", "Cambio por consola de videojuegos o tablet."),
        ("Instrumento musical", "Bateria Acustica Pearl Roadshow", "Ambos", "Bateria completa de 5 piezas con platillos y herrajes", "RD$ 35,000 - 45,000", "Cambio por guitarra de gama alta o laptop de buen rendimiento."),
        ("Instrumento musical", "Violin Acustico Cremona SV-75", "Ambos", "Violin de estudio con estuche rigido y arco de madera", "RD$ 6,000 - 9,000", "Cambio por pedales de guitarra o interfaz de audio USB."),
        ("Instrumento musical", "Microfono de Condensador Shure SM58", "Ambos", "Microfono dinamico profesional para voces en vivo o estudio", "RD$ 7,500 - 9,500", "Cambio por audifonos de estudio profesionales o controladores MIDI."),
        
        # Electrodomestico
        ("Electrodomestico", "Freidora de Aire Ninja AF101", "Ambos", "Freidora de aire de 4 cuartos de capacidad con funciones de deshidratacion", "RD$ 6,500 - 8,500", "Cambio por horno tostador moderno o licuadora de alta potencia."),
        ("Electrodomestico", "Microondas Panasonic Inverter 1.2 Cu. Ft.", "Ambos", "Microondas de acero inoxidable con tecnologia inverter", "RD$ 7,000 - 10,000", "Cambio por cafetera espresso o batidora de pedestal."),
        ("Electrodomestico", "Nevera Ejecutiva Midea 3.1 Cu. Ft.", "Ambos", "Mini nevera compacta de dos puertas ideal para oficinas o habitaciones", "RD$ 12,000 - 16,000", "Cambio por televisor inteligente de 32-43 pulgadas."),
        ("Electrodomestico", "Licuadora Oster Reversible 2 Velocidades", "Ambos", "Licuadora clasica con jarra de vidrio y motor reversible de gran potencia", "RD$ 4,500 - 6,000", "Cambio por freidora de aire basica o sandwichera."),
        ("Electrodomestico", "Aspiradora Robot Xiaomi Robot Vacuum-Mop", "Ambos", "Aspiradora y fregadora inteligente con control por app y mapeo laser", "RD$ 15,000 - 22,000", "Cambio por smartphone gama media o herramientas electricas."),

        # Electronico
        ("Electronico", "Televisor Smart TV TCL 55\" 4K", "Ambos", "Pantalla inteligente con Google TV y soporte HDR", "RD$ 22,000 - 28,000", "Cambio por consola PlayStation 4 Pro, Nintendo Switch o computadora de escritorio."),
        ("Electronico", "Bocina Inteligente Amazon Echo Dot 5ta Gen", "Ambos", "Asistente inteligente Alexa con sonido mejorado e indicador LED", "RD$ 3,000 - 4,500", "Cambio por audifonos inalambricos basicos o cargador portatil de alta capacidad."),
        ("Electronico", "Camara Reflex Canon EOS Rebel T7", "Ambos", "Camara DSLR para fotografia y video con lente 18-55mm y conectividad Wi-Fi", "RD$ 25,000 - 32,000", "Cambio por laptop para edicion de fotos o iPad."),
        ("Electronico", "Proyector Portatil Anker Nebula Capsule", "Ambos", "Proyector inteligente de bolsillo con altavoz integrado y Android TV", "RD$ 18,000 - 24,000", "Cambio por consola Nintendo Switch o tableta grafica de gama media."),
        ("Electronico", "Audifonos Inalambricos Sony WH-1000XM4", "Ambos", "Auriculares Over-Ear con cancelacion activa de ruido lider en el mercado", "RD$ 15,000 - 20,000", "Cambio por Apple Watch o monitor gaming de 144Hz."),

        # juegos
        ("juegos", "Consola Nintendo Switch OLED", "Ambos", "Consola hibrida con pantalla OLED de 7 pulgadas y 64GB de almacenamiento", "RD$ 18,000 - 22,000", "Cambio por iPad, smartphone equivalente o bicicleta montañera."),
        ("juegos", "Juego de Mesa Monopolio Dominicano", "Ambos", "Version especial de Monopolio con lugares y monumentos de RD", "RD$ 1,500 - 2,500", "Cambio por otros juegos de mesa como Catan, Dixit o cartas de coleccion."),
        ("juegos", "PlayStation 5 Slim 1TB", "Ambos", "Consola de ultima generacion con lector de discos y diseño compacto", "RD$ 32,000 - 38,000", "Cambio por laptop gamer, iPhone de generaciones recientes o scooter electrico."),
        ("juegos", "Figura Coleccionable Funko Pop Rara (Autografiada)", "Ambos", "Figura de coleccion especial en su caja protectora", "RD$ 4,000 - 8,000", "Cambio por otros Funko Pops raros, videojuegos o tarjetas de regalo."),
        ("juegos", "Set LEGO Star Wars Halcon Milenario", "Ambos", "Set de construccion detallado con mas de 1,300 piezas", "RD$ 9,000 - 13,000", "Cambio por tableta Android o instrumentos musicales de estudio."),

        # Mueble (hogar)
        ("Mueble (hogar)", "Juego de Comedor de Madera de Caoba (4 sillas)", "Ambos", "Mueble clasico tallado a mano con tapizado en tela fina", "RD$ 25,000 - 35,000", "Cambio por electrodomesticos grandes (estufa, lavadora) o juego de muebles de sala."),
        ("Mueble (hogar)", "Sofá Cama Tapizado Gris de 3 Plazas", "Ambos", "Sofá convertible en cama de dos plazas con estructura metalica robusta", "RD$ 18,000 - 24,000", "Cambio por televisor Smart TV moderno o consola de videojuegos."),
        ("Mueble (hogar)", "Escritorio Minimalista en L", "Ambos", "Escritorio de oficina con base de metal y tope de madera industrial", "RD$ 6,000 - 9,500", "Cambio por silla ergonomica de oficina o monitor para PC."),
        ("Mueble (hogar)", "Estante Modular para Libros de 5 Niveles", "Ambos", "Librero de madera melamina wengue de facil armado", "RD$ 4,500 - 7,000", "Cambio por decoraciones para el hogar, lampara de pie o mesa de noche."),
        ("Mueble (hogar)", "Gabinete Auxiliar de Cocina con Ruedas", "Ambos", "Mueble organizador con tope de madera para microondas y gavetas", "RD$ 5,000 - 8,000", "Cambio por utensilios de cocina de alta gama o electrodomestico pequeño."),

        # Vehiculos
        ("Vehiculos", "Bicicleta Montañera Trek Marlin 5", "Ambos", "Bicicleta de montaña aro 29 con suspension delantera y frenos de disco hidraulicos", "RD$ 24,000 - 30,000", "Cambio por iPhone de generacion reciente, consola PS5 o laptop."),
        ("Vehiculos", "Scooter Electrico Xiaomi Essential", "Ambos", "Patineta electrica urbana con velocidad max. de 20 km/h y autonomia de 20 km", "RD$ 15,000 - 22,000", "Cambio por bicicleta de ruta o tableta iPad."),
        ("Vehiculos", "Casco para Motocicleta LS2 Stream EVO", "Ambos", "Casco integral certificado DOT con visor solar interno y pinlock", "RD$ 6,500 - 9,000", "Cambio por chaquetas para motorista, guantes de proteccion o accesorios de audio para moto."),
        ("Vehiculos", "Pasola Yamaha Jog Artistica (Usada)", "Ambos", "Pasola clasica con papeles al dia y motor en excelentes condiciones", "RD$ 35,000 - 45,000", "Cambio por laptop de alta gama para diseño, camara DSLR con lentes o pago en efectivo."),
        ("Vehiculos", "Bomba de Aire Electrica para Neumaticos Xiaomi", "Ambos", "Compresor portatil inteligente para autos, motos y bicicletas con bateria", "RD$ 2,800 - 4,000", "Cambio por herramientas de mano o cargador de bateria de carro."),

        # Herramienta
        ("Herramienta", "Taladro Inalambrico DeWalt 20V Max", "Ambos", "Taladro percutor con dos baterias de litio, cargador y maletin de transporte", "RD$ 9,000 - 13,000", "Cambio por sierra circular, lijadora de banda u otras herramientas equivalentes."),
        ("Herramienta", "Caja de Herramientas Mecanicas Stanley (150 piezas)", "Ambos", "Maletin de llaves de vaso (dados), llaves fijas, destornilladores y accesorios", "RD$ 8,000 - 11,500", "Cambio por multimetro profesional o prensa de banco."),
        ("Herramienta", "Hidrolavadora a Presion Kärcher K2", "Ambos", "Lavadora a presion de agua fria para limpieza del hogar y automoviles", "RD$ 7,500 - 10,000", "Cambio por podadora de cesped o aspiradora industrial."),
        ("Herramienta", "Esmeriladora Angular Bosch 4 1/2\"", "Ambos", "Esmeriladora potente para corte y desbaste de metal con mango auxiliar", "RD$ 4,500 - 6,500", "Cambio por juegos de brocas profesionales o prensa de carpinteria."),
        ("Herramienta", "Juego de Destornilladores de Precision iFixit Pro Tech", "Ambos", "Kit de herramientas de reparacion para smartphones, laptops y consolas", "RD$ 3,500 - 5,000", "Cambio por accesorios electronicos, memorias SSD o discos duros externos."),

        # Joya
        ("Joya", "Reloj para Caballero Seiko 5 Automatico", "Ambos", "Reloj analógico clasico de acero inoxidable con movimiento automatico japones", "RD$ 12,000 - 18,000", "Cambio por smartphone gama media o reloj inteligente Garmin/Apple Watch."),
        ("Joya", "Cadena de Plata Ley .925 (Diseño Fígaro)", "Ambos", "Cadena de plata italiana para cuello de 55cm de longitud", "RD$ 4,500 - 7,000", "Cambio por anillo de plata o locion original de marca premium."),
        ("Joya", "Anillo de Oro Amarillo de 14k con Circonias", "Ambos", "Anillo clasico elegante con piedra central y detalles tallados", "RD$ 15,000 - 22,000", "Cambio por electrodomesticos pequeños o tecnologia equivalente."),
        ("Joya", "Reloj Inteligente Apple Watch Series 8 GPS", "Ambos", "Smartwatch con caja de aluminio de 41mm y correa deportiva", "RD$ 16,000 - 22,000", "Cambio por tableta iPad, Nintendo Switch o camara GoPro."),
        ("Joya", "Aretes de Oro de 10k Tipo Broquel para Bebé", "Ambos", "Pendientes pequeños y seguros de oro con cierre de rosca", "RD$ 3,500 - 5,500", "Cambio por articulos infantiles nuevos (coche de bebe, silla para auto)."),

        # Clases/Lecciones
        ("Clases/Lecciones", "Curso Completo de Ingles para Negocios (1 mes)", "Intercambio", "Lecciones personalizadas de conversacion y redaccion formal corporativa", "RD$ 5,000 - 8,000", "Cambio por servicios de diseño grafico (creacion de logo) o mentoria en marketing."),
        ("Clases/Lecciones", "Clases de Guitarra o Piano a Domicilio (4 sesiones)", "Intercambio", "Clases practicas individuales para principiantes de 1 hora de duracion", "RD$ 4,000 - 6,000", "Cambio por afinacion de instrumentos, libros de musica o traduccion de documentos."),
        ("Clases/Lecciones", "Tutoria Universitaria de Calculo y Algebra", "Intercambio", "Preparacion intensiva para examenes parciales o finales de ingenieria/ciencias", "RD$ 3,000 - 5,000", "Cambio por libros de texto universitarios, calculadora cientifica avanzada o cafe premium."),
        ("Clases/Lecciones", "Curso de Fotografia Digital y Edicion Basica (Online)", "Intercambio", "Lecciones en video y tutorias semanales sobre manejo de camara y Lightroom", "RD$ 4,500 - 7,500", "Cambio por accesorios de camara (tripode, filtros) o licencias de software."),
        ("Clases/Lecciones", "Clase de Cocina Italiana y Reposteria (Grupal)", "Intercambio", "Taller practico presencial donde aprenderas a hacer pasta fresca y tiramisu", "RD$ 3,500 - 5,000", "Cambio por insumos de reposteria, ingredientes importados o utensilios de cocina."),

        # Monetario
        ("Monetario", "Tarjeta de Regalo Amazon US$ 100", "Ambos", "Codigo de tarjeta de regalo digital canjeable en Amazon USA", "RD$ 5,800 - 6,200", "Cambio por articulos fisicos de igual valor o pago directo."),
        ("Monetario", "Gift Card Netflix de RD$ 2,000", "Ambos", "Tarjeta fisica prepagada para recargar saldo de suscripcion de Netflix", "RD$ 1,800 - 2,000", "Cambio por videojuegos de Nintendo/PlayStation o accesorios moviles."),
        ("Monetario", "Bono de Compra de Supermercado Nacional de RD$ 5,000", "Ambos", "Bono fisico al portador valido para canjear en alimentos y mercancia general", "RD$ 4,800 - 5,000", "Cambio por articulos para el hogar, herramientas o electrodomesticos pequeños."),
        ("Monetario", "Cupon de Descuento Especial en Tienda de Ropa (50% Off)", "Ambos", "Cupon transferible exclusivo aplicable en tienda de marca internacional", "RD$ 1,000 - 2,500", "Cambio por perfumes, accesorios de vestir o libros."),
        ("Monetario", "Tarjeta de Regalo PlayStation Store US$ 50", "Ambos", "Tarjeta digital para comprar juegos, expansiones o suscripcion de PS Plus", "RD$ 2,800 - 3,100", "Cambio por accesorios gaming (mouse, teclado) o tarjetas de regalo equivalentes."),

        # Adulto
        ("Adulto", "Perfume Carolina Herrera Good Girl 80ml", "Ambos", "Eau de Parfum original para dama con el frasco iconico en forma de tacon", "RD$ 6,500 - 8,500", "Cambio por otros perfumes originales para caballeros o maquillaje de alta gama."),
        ("Adulto", "Lentes de Sol Ray-Ban Aviator Clasicos", "Ambos", "Lentes de sol unisex originales con montura dorada y estuche original", "RD$ 8,000 - 11,000", "Cambio por smartwatch basico o audifonos bluetooth de marca."),
        ("Adulto", "Set de Cuidado Facial de Lujo Clinique", "Ambos", "Kit de 3 pasos para el cuidado y limpieza profunda de la piel del rostro", "RD$ 4,500 - 6,500", "Cambio por planchas de pelo profesionales, secadores o maquillaje premium."),
        ("Adulto", "Vaporizador Recargable Smok Nord 5 Kit", "Ambos", "Dispositivo de vapeo avanzado con pantalla digital y control de flujo de aire", "RD$ 3,000 - 4,500", "Cambio por esencias premium, cargador de baterias de litio o accesorios tecnologicos."),

        # Cuidado personal
        ("Cuidado personal", "Secador de Pelo Profesional BaBylissPRO Nano Titanium", "Ambos", "Secador de alta potencia y durabilidad preferido en salones de belleza", "RD$ 6,000 - 8,500", "Cambio por plancha de pelo de la misma marca o set de cosmeticos profesionales."),
        ("Cuidado personal", "Rasuradora Electrica Recargable Philips Norelco Series 5000", "Ambos", "Afeitadora rotativa para uso seco y humedo con cortapatillas desplegable", "RD$ 4,500 - 6,500", "Cambio por perfumes de caballero originales o recortadora de barba avanzada."),
        ("Cuidado personal", "Plancha de Pelo Remington Keratin Therapy", "Ambos", "Plancha con placas de ceramica infundidas con microacondicionadores de queratina", "RD$ 3,500 - 5,000", "Cambio por cepillo secador giratorio o set de manicura profesional."),
        ("Cuidado personal", "Espejo de Maquillaje con Luces LED y Aumento 10x", "Ambos", "Espejo de mesa giratorio recargable con brillo ajustable", "RD$ 2,000 - 3,500", "Cambio por organizador de cosmeticos de acrilico o accesorios de belleza."),
        ("Cuidado personal", "Cepillo de Dientes Electrico Oral-B Pro 1000", "Ambos", "Cepillo dental electrico recargable con sensor de presion inteligente", "RD$ 3,500 - 5,000", "Cambio por irrigador bucal portatil u otros articulos de cuidado bucal."),

        # Decoraciones
        ("Decoraciones", "Cuadro al Oleo Paisaje Dominicano", "Ambos", "Pintura original firmada por artista local que representa un atardecer en el campo", "RD$ 5,000 - 9,000", "Cambio por lampara de mesa decorativa, alfombras de sala o espejos de pared elegantes."),
        ("Decoraciones", "Espejo de Pared Redondo con Marco de Bambu", "Ambos", "Espejo decorativo de diseño nordico/boho ideal para salas o recibidores", "RD$ 3,500 - 5,000", "Cambio por macetas de ceramica grandes para interiores o cojines decorativos."),
        ("Decoraciones", "Alfombra de Sala Pelo Corto Geometrica 5x7 Ft", "Ambos", "Alfombra moderna y lavable de tonos neutros (gris y blanco)", "RD$ 6,000 - 9,000", "Cambio por mesitas auxiliares de madera o cortinas de sala premium."),
        ("Decoraciones", "Juego de 3 Repisas Flotantes de Madera Rustica", "Ambos", "Estantes decorativos de facil instalacion con soportes ocultos de metal", "RD$ 2,500 - 4,000", "Cambio por portarretratos multiples, velas aromaticas de lujo o plantas artificiales."),
        ("Decoraciones", "Lampara de Pie Moderna en Tripode de Madera", "Ambos", "Lampara de salon con pantalla de tela blanca y estructura de tres patas", "RD$ 4,500 - 6,500", "Cambio por cuadros decorativos o jarrones de ceramica."),

        # Deportes
        ("Deportes", "Mancuernas Ajustables de 24 kg cada una (Par)", "Ambos", "Pesas de sistema selector rapido ideales para gimnasio en casa", "RD$ 15,000 - 22,000", "Cambio por bicicleta estatica, maquina eliptica o barra de levantamiento."),
        ("Deportes", "Raqueta de Tenis Babolat Pure Drive (Aero)", "Ambos", "Raqueta profesional de grafito ligera para potencia y efectos", "RD$ 10,000 - 14,000", "Cambio por raqueta de padel de gama media-alta o bolso para tenis."),
        ("Deportes", "Bolsa de Boxeo Everlast", "Ambos", "Saco de boxeo de cuero sintetico relleno de 80 lbs para entrenamiento de impacto", "RD$ 5,500 - 8,000", "Cambio por guantes de boxeo de piel y vendas, o kits de crossfit."),
        ("Deportes", "Smartband Reloj Deportivo Garmin Forerunner 55", "Ambos", "Reloj con GPS para running con sensor de ritmo cardiaco y metricas avanzadas", "RD$ 11,000 - 15,000", "Cambio por tableta grafica o monitor gaming."),
        ("Deportes", "Pelota de Baloncesto Wilson Evolution Oficial (Size 7)", "Ambos", "Balon oficial de cuero compuesto favorito para canchas techadas", "RD$ 4,500 - 6,000", "Cambio por pelota de futbol de alta gama o indumentaria deportiva original."),

        # Hogar
        ("Hogar", "Juego de Sabanas de Algodon Egipcio King Size", "Ambos", "Juego de 400 hilos super suaves que incluye sabana bajera, encimera y fundas", "RD$ 4,500 - 7,000", "Cambio por edredon reversible, toallas de baño de lujo o almohadas ortopedicas."),
        ("Hogar", "Vajilla de Ceramica Moderna de 16 Piezas", "Ambos", "Servicio completo para 4 personas con platos llanos, hondos, de postre y tazas", "RD$ 3,800 - 5,500", "Cambio por juego de cubiertos de acero inoxidable de 24 piezas o copas de cristal."),
        ("Hogar", "Juego de Ollas de Cocina Tramontina (7 Piezas)", "Ambos", "Bateria de cocina de aluminio con revestimiento antiadherente interno", "RD$ 6,500 - 9,000", "Cambio por extractor de grasa, sarten de hierro fundido de lujo o tostadora de 4 panes."),
        ("Hogar", "Organizador de Zapatos de Madera para Armario", "Ambos", "Zapatera modular para hasta 30 pares de zapatos con diseño vertical", "RD$ 3,000 - 5,000", "Cambio por organizador de ropa, percheros de pared o cestas de lavanderia."),
        ("Hogar", "Purificador de Aire Portatil Holmes con Filtro HEPA", "Ambos", "Purificador para habitaciones pequeñas con filtro que elimina alergenos y olores", "RD$ 4,000 - 6,500", "Cambio por ventilador de torre con control remoto o difusor ultrasonico grande."),

        # Jardín
        ("Jardín", "Podadora de Cesped de Gasolina Craftsman 140cc", "Ambos", "Maquina podadora potente con bolsa recolectora trasera y ajuste de altura", "RD$ 16,000 - 22,000", "Cambio por soplador de hojas inalambrico o hidrolavadora."),
        ("Jardín", "Juego de Comedor de Exterior en Rattan Sintetico", "Ambos", "Mesa redonda con cristal y cuatro sillas comodas resistentes a la intemperie", "RD$ 22,000 - 30,000", "Cambio por barbacoa de carbon grande o muebles de sala de interior."),
        ("Jardín", "Barbacoa de Carbon Weber Original Kettle 18\"", "Ambos", "Parrilla de carbon clasica con tapa esmaltada y recolector de ceniza", "RD$ 8,500 - 12,000", "Cambio por herramientas de jardin electricas (recortadora de setos) u ollas de cocina."),
        ("Jardín", "Kit de Herramientas de Jardineria de 10 piezas", "Ambos", "Incluye palas, rastrillo, tijeras de podar, rociador y guantes con estuche", "RD$ 2,500 - 3,500", "Cambio por macetas de terracota decoradas o fertilizantes organicos premium."),
        ("Jardín", "Manguera Expandible para Jardinar de 100 pies", "Ambos", "Manguera ligera que se expande con la presion de agua y boquilla de 9 funciones", "RD$ 1,800 - 2,800", "Cambio por luces LED decorativas solares para jardin o aspersor giratorio."),

        # Teléfonos
        ("Teléfonos", "Smartphone Samsung Galaxy S22 Ultra 256GB", "Ambos", "Telefono de gama alta con S-Pen integrado, pantalla Dynamic AMOLED y excelente camara", "RD$ 25,000 - 32,000", "Cambio por computadora portatil, iPhone de similar gama o consola PlayStation 5."),
        ("Teléfonos", "iPhone 13 Pro Max 128GB (Factory Unlocked)", "Ambos", "iPhone con bateria en buen estado, color azul sierra, libre de fabrica", "RD$ 28,000 - 35,000", "Cambio por iPad Air/Pro reciente, laptop gamer o MacBook Air."),
        ("Teléfonos", "Audifonos Inalambricos Apple AirPods Pro 2da Gen", "Ambos", "Auriculares con cancelacion activa de ruido y estuche con carga MagSafe USB-C", "RD$ 10,000 - 13,000", "Cambio por Apple Pencil de 2da gen + cargador de pared Apple original o Nintendo Switch Lite."),
        ("Teléfonos", "Reloj Inteligente Samsung Galaxy Watch 5 Pro 45mm", "Ambos", "Smartwatch resistente con cristal de zafiro, cuerpo de titanio y gran bateria", "RD$ 12,000 - 16,000", "Cambio por audifonos inalambricos premium Sony o tableta Android."),
        ("Teléfonos", "Cargador Portatil Inalambrico Anker 10,000mAh", "Ambos", "Bateria magnetica externa portatil para iPhone e inalambricos", "RD$ 3,000 - 4,200", "Cambio por fundas originales, cargadores de escritorio rapidos u otros accesorios."),

        # Niños
        ("Niños", "Carro de Juguete Electrico Montable Jeep 12V", "Ambos", "Vehiculo montable con control remoto para padres, luces LED y reproductor MP3", "RD$ 12,000 - 18,000", "Cambio por consola Nintendo Switch Lite, bicicleta montañera infantil o tablet."),
        ("Niños", "Bicicleta Infantil Aro 16 con Rueditas", "Ambos", "Bicicleta color azul con freno de contrapedal y protector de cadena para seguridad", "RD$ 4,500 - 6,500", "Cambio por juguetes educativos (bloques, Legos) o ropa infantil de marca."),
        ("Niños", "Coche de Bebe Chicco Bravo Quick-Fold", "Ambos", "Cochecito plegable con portavasos, capota extensible y gran cesta portaobjetos", "RD$ 10,000 - 15,000", "Cambio por cuna de viaje portatil, extractor de leche electrico o silla para el coche."),
        ("Niños", "Consola de Videojuegos Portatil retro Powkiddy RGB20S", "Ambos", "Dispositivo portatil con miles de juegos clasicos de NES, SNES, GBA y PS1", "RD$ 3,500 - 5,000", "Cambio por juegos de Nintendo Switch o accesorios infantiles."),

        # Antigüedades
        ("Antigüedades", "Maquina de Escribir Manual Remington Rand (1950s)", "Ambos", "Maquina de escribir vintage funcional en color gris con su estuche original de madera", "RD$ 8,000 - 14,000", "Cambio por camaras de fotos analogicas antiguas, tocadiscos o discos de vinilo de rock clasicos."),
        ("Antigüedades", "Camara Fotografica de Fuelle Kodak No. 1A (Años 1920)", "Ambos", "Camara de coleccion historica ideal para decoracion de estudios o bibliotecas", "RD$ 5,000 - 9,000", "Cambio por libros antiguos ilustrados, monedas de coleccion o pinturas antiguas."),
        ("Antigüedades", "Tocadiscos Maleta de Vinilo Victrola", "Ambos", "Reproductor de discos de vinilo clasico en maletin portatil con Bluetooth", "RD$ 4,500 - 7,000", "Cambio por albumes de vinilo originales, audifonos over-ear o decoraciones vintage."),
        ("Antigüedades", "Lampara de Escritorio Vintage Estilo Banquero", "Ambos", "Lampara clasica de oficina con base de laton y tulipa de cristal verde esmeralda", "RD$ 3,500 - 5,500", "Cambio por sujetalibros de bronce, plumas estilograficas o encendedores antiguos."),
        ("Antigüedades", "Moneda de Plata de 1 Peso Dominicano de 1897", "Ambos", "Moneda historica de plata dominicana de alta pureza en excelente estado de conservacion", "RD$ 6,000 - 12,000", "Cambio por relojes mecanicos antiguos u otras piezas numismaticas valiosas."),

        # Niñas
        ("Niñas", "Casa de Muñecas de Madera de 3 Pisos KidKraft", "Ambos", "Mansion de madera detallada para muñecas de 12 pulgadas con muebles incluidos", "RD$ 9,000 - 14,000", "Cambio por tablet infantil, bicicleta de niña o set de patines."),
        ("Niñas", "Bicicleta de Niña Aro 16 Huffy (Color Rosado)", "Ambos", "Bicicleta con cesta delantera, serpentinas en el manillar y ruedas de aprendizaje", "RD$ 4,500 - 6,500", "Cambio por muñecas de coleccion, bloques de construccion o ropa de marca."),
        ("Niñas", "Set de Patines Ajustables en Linea Rosados Rollerblade", "Ambos", "Patines ajustables de talla con kit de proteccion (casco, coderas y rodilleras)", "RD$ 3,800 - 5,500", "Cambio por mochilas escolares de marca, set de pintura o juguetes interactivos."),
        ("Niñas", "Muñeca Interactiva Baby Alive", "Ambos", "Muñeca con accesorios de comida y pañales incluidos", "RD$ 2,500 - 3,800", "Cambio por set de manualidades, plastilinas Play-Doh o juegos de mesa infantiles."),

        # Mascotas
        ("Mascotas", "Caja Transportadora Petmate Sky Kennel", "Ambos", "Transportador de viaje seguro homologado para vuelos de mascotas", "RD$ 5,000 - 8,000", "Cambio por camas ortopedicas para perros grandes o casas de exterior."),
        ("Mascotas", "Rascador Modular para Gatos de 5 Niveles", "Ambos", "Arbol rascador forrado de felpa y cuerdas de sisal para que los gatos escalen", "RD$ 4,500 - 6,800", "Cambio por fuentes de agua automaticas, alimentadores programables o arneses."),
        ("Mascotas", "Filtro de Acuario Externo Canister SunSun", "Ambos", "Filtro de alto rendimiento para acuarios de agua dulce o salada de hasta 75 galones", "RD$ 6,000 - 8,500", "Cambio por calentadores, luces LED para acuarios plantados o decoraciones de resina."),
        ("Mascotas", "Cama Ortopedica para Mascotas", "Ambos", "Cama grande con funda lavable e impermeable para la comodidad de perros senior", "RD$ 3,000 - 4,500", "Cambio por alimento para mascotas de marca premium o juguetes mordedores de alta durabilidad."),
        ("Mascotas", "Alimentador Automatico Programable", "Ambos", "Dispensador de comida seca con pantalla LCD para programar porciones y horarios", "RD$ 4,000 - 6,000", "Cambio por camaras de vigilancia de mascotas o secadoras portatiles para pelo de mascotas."),

        # Tecnología
        ("Tecnología", "Laptop HP Pavilion 15.6\" i5 8GB 512GB SSD", "Ambos", "Laptop veloz ideal para estudios, trabajo de oficina y navegacion web", "RD$ 24,000 - 30,000", "Cambio por consola de videojuegos PS5, iPhone de similar gama o iPad Air."),
        ("Tecnología", "Monitor Gaming Curvo MSI 27\" 165Hz 1ms Optix", "Ambos", "Pantalla con panel VA, resolucion FHD y tecnologia FreeSync ideal para shooters", "RD$ 13,000 - 17,500", "Cambio por tarjetas graficas (GPU) de gama media, procesador Ryzen o tablet."),
        ("Tecnología", "Memoria SSD Portatil Externa SanDisk 1TB USB 3.2", "Ambos", "Disco de estado solido ultra resistente y rapido para transferir fotos o videos", "RD$ 5,500 - 7,500", "Cambio por memorias RAM DDR4, mouse y teclado mecanico inalambricos o routers Wi-Fi mesh."),
        ("Tecnología", "Teclado Mecanico Inalambrico Keychron K2 RGB", "Ambos", "Teclado compacto del 75% con switches mecanicos Gateron marrones e intercambiables", "RD$ 6,000 - 8,500", "Cambio por audifonos para PC de buena calidad o microfonos USB para streaming."),
        ("Tecnología", "Tarjeta Grafica NVIDIA GTX 1660 Super 6GB (Usada)", "Ambos", "GPU en excelentes condiciones para juegos a 1080p", "RD$ 10,000 - 14,000", "Cambio por procesador de PC equivalente, memorias SSD de gran capacidad o consolas retro."),

        # Librería y Papelería
        ("Librería y Papelería", "Juego de Marcadores Copic Sketch (12 piezas)", "Ambos", "Rotuladores recargables con puntas de pincel altamente cotizados por ilustradores", "RD$ 4,500 - 6,000", "Cambio por libretas de dibujo de alta densidad (Moleskine), acuarelas profesionales o libros de arte."),
        ("Librería y Papelería", "Coleccion de Libros \"Cancion de Hielo y Fuego\"", "Ambos", "Novelas completas en español de George R.R. Martin en estuche especial", "RD$ 3,000 - 4,500", "Cambio por novelas de ciencia ficcion, comics de coleccion o novelas graficas."),
        ("Librería y Papelería", "Calculadora Grafica Texas Instruments TI-84 Plus CE", "Ambos", "Calculadora cientifica y grafica a color con bateria recargable para ingenierias", "RD$ 8,000 - 11,500", "Cambio por tablet economica o herramientas de precision."),
        ("Librería y Papelería", "Set de Pintura Oleo Profesional Winsor & Newton", "Ambos", "Caja de madera con tubos de pintura al oleo fina, pinceles, espatulas y disolventes", "RD$ 5,500 - 8,000", "Cambio por caballete de madera ajustable, lienzos grandes o tabletas de dibujo básicas."),

        # Damas
        ("Damas", "Cartera de Cuero Genuino Michael Kors", "Ambos", "Bolso de hombro original de cuero saffiano en perfectas condiciones con bolsa de polvo", "RD$ 9,000 - 13,000", "Cambio por perfumes importados de dama originales, billeteras de marcas de lujo o relojes de marca."),
        ("Damas", "Perfume Chanel No. 5 Eau de Parfum 100ml", "Ambos", "Perfume original y sellado de una de las marcas mas exclusivas del mundo", "RD$ 8,500 - 11,500", "Cambio por carteras originales de gama similar, joyeria de plata/oro o Apple Watch."),
        ("Damas", "Lentes de Sol de Moda Prada para Mujer", "Ambos", "Lentes de sol originales estilo ojo de gato con estuche y paño de microfibra", "RD$ 7,500 - 10,000", "Cambio por cosmeticos importados de alta gama (Anastasia, Fenty) o calzado deportivo original."),
        ("Damas", "Vestido de Noche Corto Elegante", "Ambos", "Vestido de fiesta negro de encaje fino usado una sola vez, talla M", "RD$ 4,000 - 6,500", "Cambio por zapatos de tacon elegantes de marca, chaquetas de cuero o perfumes."),
        ("Damas", "Zapatos Deportivos Nike Air Max Dawn", "Ambos", "Tenis casuales de mujer cómodos, nuevos en su caja", "RD$ 5,000 - 7,000", "Cambio por carteras cruzadas pequeñas o ropa deportiva de marca original."),

        # Caballeros
        ("Caballeros", "Perfume Dior Sauvage Eau de Toilette 100ml", "Ambos", "Perfume original y sellado para caballero con aroma fresco y amaderado", "RD$ 7,500 - 9,500", "Cambio por otros perfumes originales de caballero o audifonos bluetooth de marca."),
        ("Caballeros", "Billetera de Cuero Genuino Montblanc", "Ambos", "Billetera clásica de piel con el icónico emblema de la marca y tarjetero", "RD$ 11,000 - 16,000", "Cambio por relojes Seiko/Citizen, chaquetas deportivas de gama alta o smartband."),
        ("Caballeros", "Chaqueta de Piel Genuina para Hombre", "Ambos", "Chaqueta estilo aviador de cuero real pesado, excelente calidad, talla L", "RD$ 8,000 - 13,000", "Cambio por botas de cuero de marca, relojes analógicos o artículos de tecnología."),
        ("Caballeros", "Zapatos Deportivos Adidas Ultraboost Light", "Ambos", "Tenis de correr de alta gama con tecnología Boost, nuevos en caja", "RD$ 9,000 - 12,000", "Cambio por monitor gamer, Apple Pencil o tarjetas de regalo."),
        ("Caballeros", "Lentes de Sol Deportivos Oakley Holbrook", "Ambos", "Gafas de sol con montura negra mate y lentes polarizados prizm", "RD$ 7,000 - 9,500", "Cambio por mochilas de viaje profesionales o accesorios de PC."),

        # Oficina
        ("Oficina", "Silla de Oficina Ergonómica Sihoo M18", "Ambos", "Silla ergonómica con soporte lumbar ajustable, malla transpirable y reposacabezas", "RD$ 10,000 - 14,000", "Cambio por monitor Full HD de 24 pulgadas, escritorio para computadora o teclado mecánico."),
        ("Oficina", "Impresora Multifuncional Epson EcoTank L3210", "Ambos", "Impresora con sistema de tanques de tinta de altísimo rendimiento y escáner", "RD$ 11,000 - 14,500", "Cambio por tablet de marca o smartphone gama media-baja."),
        ("Oficina", "Plastificadora de Documentos Profesional A3", "Ambos", "Laminadora para sellar documentos e imágenes con calor y frío", "RD$ 4,500 - 6,500", "Cambio por destructora de papel compacta o suministros de oficina premium."),
        ("Oficina", "Organizador de Escritorio de Cuero Sintético", "Ambos", "Incluye tapete de escritorio, bandeja de cartas, cubilete de lápices y archivador", "RD$ 2,500 - 4,000", "Cambio por lámpara de escritorio LED con cargador inalámbrico o mouse ergonómico."),
        ("Oficina", "Destructora de Papel de Corte Cruzado", "Ambos", "Trituradora de documentos y tarjetas de crédito con ranura de alimentación rápida", "RD$ 4,000 - 6,000", "Cambio por organizadores de cables, soportes dobles de monitor o discos duros."),

        # Talentos-Servicios
        ("Talentos-Servicios", "Servicio de Diseño de Logotipo e Identidad Visual", "Intercambio", "Creación de logotipo vectorial profesional, paleta de colores y tipografías para marca", "RD$ 6,000 - 10,000", "Cambio por artículos de tecnología (discos SSD, memoria RAM), perfumes originales o lecciones de música."),
        ("Talentos-Servicios", "Servicio de Fotografía de Productos (15 fotos)", "Intercambio", "Sesión fotográfica de estudio con retoque digital ideal para páginas web o e-commerce", "RD$ 8,000 - 12,000", "Cambio por trípode profesional de video, tarjetas de memoria SDXC rápidas o discos externos."),
        ("Talentos-Servicios", "Traducción Profesional de Documentos (Inglés - Español)", "Intercambio", "Traducción precisa y formateada de hasta 15 páginas de contenido legal, comercial o académico", "RD$ 4,500 - 7,000", "Cambio por audífonos inalámbricos de marca, accesorios para el hogar o mentorías de negocios."),
        ("Talentos-Servicios", "Asesoría Contable y Preparación de Declaración Jurada IR-1", "Intercambio", "Revisión fiscal individual y preparación de la declaración de impuestos ante la DGII", "RD$ 8,000 - 12,000", "Cambio por herramientas eléctricas, electrodomésticos pequeños o cursos formativos."),
        ("Talentos-Servicios", "Sesión de Consultoría de Marketing Digital y Redes Sociales", "Intercambio", "Análisis de marca en Instagram/TikTok, auditoría de contenido y propuesta estratégica", "RD$ 5,000 - 8,000", "Cambio por libros de marketing, micrófonos de solapa inalámbricos o tarjetas de regalo.")
    ]
    
    # Write data rows
    for row_idx, item in enumerate(data, start=2):
        ws.append(item)
        row_fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
        
        for col_idx in range(1, len(item) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = item_name_font if col_idx == 2 else default_font
            cell.border = cell_border
            
            # Alignments
            if col_idx == 3: # Tipo Recomendado
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif col_idx == 5: # Valor Estimado
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Auto-adjust column widths with paddings
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Calculate maximum length
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        
        # Apply standard or custom widths to look premium
        if col_letter in ['A', 'C']: # Categoria, Tipo
            ws.column_dimensions[col_letter].width = max(max_len + 4, 18)
        elif col_letter in ['B', 'E']: # Articulo, Valor
            ws.column_dimensions[col_letter].width = max(max_len + 4, 25)
        elif col_letter in ['D', 'F']: # Descripcion, Criterios
            ws.column_dimensions[col_letter].width = 45 # Constrain description columns to wrap
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save to file in the workspace
    filepath = "Listado_Articulos_Intercambios.xlsx"
    wb.save(filepath)
    print("Success: Excel file created successfully: " + filepath)

if __name__ == "__main__":
    create_excel()
